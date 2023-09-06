
from openpyxl import load_workbook
import os

class doExcel():
    def __init__(self,filename=None,sheetId=None):

        if filename:
            self.filename=filename
            self.sheetId=sheetId
        else:
            self.filename="person.xlsx"
            self.sheetId="data1"
        self.path = os.path.join(os.path.abspath(os.path.dirname(__file__)), self.filename)
        self.data=self.getData()
        print("数据来源：%s" % self.path)
    def getData(self):
        wb = load_workbook(self.path)
        sheet = wb[self.sheetId]

        test_data = []
        for i in range(1, sheet.max_row):
            sub_data = {}
            sub_data['caseId'] = sheet.cell(i + 1, 1).value
            sub_data['interface'] = sheet.cell(i + 1, 2).value
            sub_data['url'] = sheet.cell(i + 1, 3).value
            sub_data['isRun'] = sheet.cell(i + 1, 4).value
            sub_data['method'] = sheet.cell(i + 1, 5).value
            sub_data['data'] = sheet.cell(i + 1, 6).value
            sub_data['expected'] = sheet.cell(i + 1, 7).value
            sub_data['acturally'] = sheet.cell(i + 1, 8).value  # 验证目的
            test_data.append(sub_data)
        return test_data

if __name__=='__main__':
    a=doExcel().getData()
    print(a)
